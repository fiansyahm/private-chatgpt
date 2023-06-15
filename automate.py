import shutil
import time
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import NoSuchElementException
from selenium.webdriver.common.action_chains import ActionChains
import os
import pandas
from openpyxl import Workbook

from flask import Flask, request, render_template,send_file
import pandas as pd

app = Flask(__name__)

def click_selenium(driver,xfullpath,search_by,send_keys=None):
 # tunggu hingga elemen yang ingin diklik terlihat dan dapat diklik
    wait = WebDriverWait(driver, 10)
    if search_by == "xpath":
        by_for=By.XPATH
    elif search_by == "id":
        by_for=By.ID
    elif search_by == "css":
        by_for=By.CSS_SELECTOR
    click_elem = wait.until(EC.element_to_be_clickable((by_for, xfullpath)))
    
    if send_keys == "clear": 
        click_elem.clear()
    elif send_keys == "special": 
        ActionChains(driver).click(click_elem).perform()
    elif send_keys != None:
        click_elem.send_keys(send_keys)
    else:
        # klik elemen yang sudah terlihat dan dapat diklik
        click_elem.click() 
    return click_elem

def get_selenium(driver,xfullpath,search_by):
 # tunggu hingga elemen yang ingin diklik terlihat dan dapat diklik
    wait = WebDriverWait(driver, 10)
    if search_by == "xpath":
        by_for=By.XPATH
    elif search_by == "id":
        by_for=By.ID
    elif search_by == "css":
        by_for=By.CSS_SELECTOR
    click_elem = wait.until(EC.element_to_be_clickable((by_for, xfullpath)))
    return click_elem.text

def run_selenium(csv_file,email,password,timer):
    # Specify the folder path to delete
    folder_to_delete = "output"

    # Delete the folder if it exists
    if os.path.exists(folder_to_delete):
        shutil.rmtree(folder_to_delete)
        print("Folder deleted:", folder_to_delete)
    else:
        print("Folder does not exist:", folder_to_delete)

    # Specify the folder path to create
    folder_to_create = "output"

    # Create the folder if it doesn't exist
    if not os.path.exists(folder_to_create):
        os.makedirs(folder_to_create)
        print("Folder created:", folder_to_create)
    else:
        print("Folder already exists:", folder_to_create)


    isiexcel = pandas.read_csv(csv_file)
    myexcelcontain = isiexcel.values
    # Saving the Excel file
    workbook = Workbook()
    worksheet = workbook.active

    row = 1
    col = 1

    worksheet.cell(row=row, column=col, value="prompt")
    worksheet.cell(row=row, column=col+1, value="output")
    row += 1
    
    chromedriver_path = r"chromedriver.exe"  # Update with the actual path

    if os.path.exists(chromedriver_path):
        print("The file exists")
    else:
        print("The file does not exist")
        chromedriver_path = r"chromedriver.exe"  # Update with the actual path

    # Initialize the WebDriver with the chromedriver path
    driver = webdriver.Chrome(executable_path=chromedriver_path)

    # Membuka halaman web "https://crewdible.com/seller-app"
    driver.get("https://ai.undrctrl.id/#/chat")

    driver.maximize_window()

    xfullpath='/html/body/div[1]/div/div[1]/div/div/div/aside/div[1]/div/footer/div/div/div[2]/a/span/div/h2'
    click_selenium(driver,xfullpath,"xpath")
        
    # logins
    #    email
    xfullpath='/html/body/div[2]/div/div/div[1]/div/div[3]/div/div/div[2]/div[1]/div[1]/div/input'
    click_selenium(driver,xfullpath,"xpath",email)

    #    password
    xfullpath='/html/body/div[2]/div/div/div[1]/div/div[3]/div/div/div[2]/div[2]/div[1]/div/input'
    click_selenium(driver,xfullpath,"xpath",password)

    time.sleep(10)

    xfullpath='/html/body/div[2]/div/div/div[1]/div/div[3]/div/div/div[2]/div[2]/div[1]/div/input'
    click_selenium(driver,xfullpath,"xpath","\n")


    for i in range(len(myexcelcontain)):
        driver.get("https://ai.undrctrl.id/#/chat")

        xfullpath='//*[@id="app"]/div/div[1]/div/div/div/aside/div[1]/div/main/div[2]/div/div[1]/div/div/div/div/div/a/div[2]/button[2]'
        click_selenium(driver,xfullpath,"xpath")

        time.sleep(2)
    
        search_keyword = "Confirm"
        script = """
        var buttons = document.getElementsByTagName('button');
        for (var i = 0; i < buttons.length; i++) {{
        if (buttons[i].textContent.includes('{0}')) {{
            buttons[i].click();
            break;
        }}
        }}
        """.format(search_keyword)

        driver.execute_script(script)

        time.sleep(5)

        prompt = myexcelcontain[i][0].replace('\n', ' ')
        try:
            topic = myexcelcontain[i][1].replace('\n', ' ')
        except:
            topic = ""
        filename = myexcelcontain[i][2].replace('\n', ' ')
        print("prompt: ", prompt)
        print("topic: ", topic)
        print("filename: ", filename)

        time.sleep(10)

        xfullpath='//*[@id="app"]/div/div[1]/div/div/div/div/div/div/footer/div/div/div[2]/div/div/div/div[1]/div[1]/textarea'
        textarea=f"{prompt} {topic}"

        click_selenium(driver,xfullpath,"xpath",textarea+"\n")

        time.sleep(timer)
        
        xfullpath='/html/body/div[1]/div/div[1]/div/div/div/div/div/div/main/div/div/div/div/div/div[2]/div[2]/div/div[2]/button[2]'
        click_selenium(driver,xfullpath,"xpath")

        xfullpath='/html/body/div[3]/div/div/div[2]'
        click_selenium(driver,xfullpath,"xpath")

        driver.get('https://www.memonotepad.com/')
        time.sleep(2)
        xfullpath='/html/body/div[4]/div/div[2]/div'
        click_selenium(driver,xfullpath,"xpath",'clear')
        click_selenium(driver,xfullpath,"xpath",(Keys.CONTROL, 'v'))
        time.sleep(2)
        mydata=get_selenium(driver,xfullpath,"xpath")
        print(mydata)

        worksheet.cell(row=row, column=col, value=prompt)
        worksheet.cell(row=row, column=col+1, value=mydata)
        row += 1

        folder_path = "output"

        # Specify the file path
        file_path = os.path.join(folder_path, f"{filename}.txt")

        # Create the text file
        with open(file_path, "w") as file:
            file.write(mydata)

        time.sleep(2)
        
    driver.close()
    workbook.save('output.xlsx')
    return workbook

@app.route('/', methods=['GET', 'POST'])
def index():
    if request.method == 'POST':
        csv_file = request.files["csv_file"]
        email=request.form.get('email')
        password=request.form.get('password')
        timer=int(request.form.get('timer'))
        workbook=run_selenium(csv_file,email,password,timer)
        return render_template('index.html', excel_data=workbook)
    return render_template('index.html')

@app.route('/download_excel')
def download_excel():
    return send_file('output.xlsx', as_attachment=True)

if __name__ == '__main__':
    app.run(debug=True)
